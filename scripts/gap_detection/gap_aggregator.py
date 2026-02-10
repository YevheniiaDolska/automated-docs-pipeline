#!/usr/bin/env python3
"""
Gap Aggregator

Агрегирует данные из всех источников в единый Excel/CSV отчёт:
- Code changes (git diff)
- Community topics (RSS)
- Search analytics (Algolia)

Выход: Excel файл для review и анализа Claude Code.
"""

import csv
import json
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

# Попытка импортировать openpyxl для Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .code_analyzer import CodeChangeAnalyzer, AnalysisResult as CodeResult
from .community_collector import CommunityCollector, CollectionResult as CommunityResult
from .algolia_parser import AlgoliaAnalytics, AlgoliaResult


@dataclass
class DocumentationGap:
    """Единый формат для документационного гэпа."""
    id: str
    title: str
    description: str
    source: str  # 'code', 'community', 'search'
    category: str
    suggested_doc_type: str  # 'tutorial', 'how-to', 'concept', 'reference', 'troubleshooting'
    priority: str  # 'high', 'medium', 'low'
    frequency: int = 1  # Как часто встречается
    keywords: list[str] = field(default_factory=list)
    related_files: list[str] = field(default_factory=list)
    sample_queries: list[str] = field(default_factory=list)
    action_required: str = ''
    status: str = 'new'  # 'new', 'in_progress', 'done', 'wont_fix'
    created_at: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class AggregatedReport:
    """Полный отчёт по документационным гэпам."""
    gaps: list[DocumentationGap] = field(default_factory=list)
    summary: dict = field(default_factory=dict)
    sources_analyzed: list[str] = field(default_factory=list)
    generated_at: str = field(default_factory=lambda: datetime.now().isoformat())


class GapAggregator:
    """Агрегирует данные из всех источников в единый отчёт."""

    def __init__(self, output_dir: str = './reports'):
        """
        Args:
            output_dir: Директория для сохранения отчётов
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def aggregate_all(
        self,
        code_result: Optional[CodeResult] = None,
        community_result: Optional[CommunityResult] = None,
        algolia_result: Optional[AlgoliaResult] = None,
    ) -> AggregatedReport:
        """
        Агрегирует данные из всех источников.

        Args:
            code_result: Результат анализа кода
            community_result: Результат сбора community
            algolia_result: Результат анализа Algolia

        Returns:
            AggregatedReport со всеми гэпами
        """
        report = AggregatedReport()
        gap_id = 0

        # Обрабатываем code changes
        if code_result:
            report.sources_analyzed.append('code_changes')
            for change in code_result.changes:
                gap_id += 1
                gap = DocumentationGap(
                    id=f'CODE-{gap_id:04d}',
                    title=f'{change.category.replace("_", " ").title()}: {self._truncate(change.description, 50)}',
                    description=change.doc_suggestion,
                    source='code',
                    category=change.category,
                    suggested_doc_type=self._map_code_to_doc_type(change.category),
                    priority=change.priority,
                    related_files=[change.file_path],
                    action_required=change.doc_suggestion,
                )
                report.gaps.append(gap)

        # Обрабатываем community topics
        if community_result:
            report.sources_analyzed.append('community')
            for suggestion in community_result.suggested_docs:
                gap_id += 1
                gap = DocumentationGap(
                    id=f'COMM-{gap_id:04d}',
                    title=f'{suggestion["category"].title()}: {suggestion["topic"]}',
                    description=f'Frequently asked topic ({suggestion["frequency"]} questions)',
                    source='community',
                    category=suggestion['category'],
                    suggested_doc_type=suggestion['suggested_doc_type'],
                    priority=suggestion['priority'],
                    frequency=suggestion['frequency'],
                    keywords=suggestion['keywords'],
                    sample_queries=suggestion['sample_questions'],
                    action_required=f'Create {suggestion["suggested_doc_type"]} covering: {", ".join(suggestion["keywords"][:3])}',
                )
                report.gaps.append(gap)

        # Обрабатываем search analytics
        if algolia_result:
            report.sources_analyzed.append('search_analytics')
            for suggestion in algolia_result.suggested_docs:
                gap_id += 1
                gap = DocumentationGap(
                    id=f'SRCH-{gap_id:04d}',
                    title=f'Search: "{suggestion["query"]}"',
                    description=f'{suggestion["reason"].replace("_", " ").title()} - {suggestion["search_count"]} searches',
                    source='search',
                    category=suggestion['category'],
                    suggested_doc_type=suggestion['suggested_doc_type'],
                    priority=suggestion['priority'],
                    frequency=suggestion['search_count'],
                    sample_queries=[suggestion['query']],
                    action_required=suggestion['action'],
                )
                report.gaps.append(gap)

        # Дедуплицируем похожие гэпы
        report.gaps = self._deduplicate_gaps(report.gaps)

        # Сортируем по приоритету и частоте
        priority_order = {'high': 0, 'medium': 1, 'low': 2}
        report.gaps.sort(key=lambda g: (priority_order.get(g.priority, 2), -g.frequency))

        # Генерируем summary
        report.summary = self._generate_summary(report)

        return report

    def run_full_analysis(
        self,
        repo_path: str = '.',
        since_days: int = 7,
        algolia_csv: Optional[str] = None,
        algolia_json: Optional[str] = None,
    ) -> AggregatedReport:
        """
        Запускает полный анализ из всех источников.

        Args:
            repo_path: Путь к git репозиторию
            since_days: Анализировать коммиты за последние N дней
            algolia_csv: Путь к CSV экспорту Algolia
            algolia_json: Путь к JSON с данными Algolia

        Returns:
            AggregatedReport
        """
        # Анализ кода
        print("Analyzing code changes...")
        code_analyzer = CodeChangeAnalyzer(repo_path)
        code_result = code_analyzer.analyze_commits(since=f'{since_days} days ago')
        print(f"  Found {len(code_result.changes)} code changes")

        # Сбор community
        print("Collecting community topics...")
        community_collector = CommunityCollector()
        community_result = community_collector.collect_all(limit_per_feed=50)
        print(f"  Found {len(community_result.topics)} topics, {len(community_result.suggested_docs)} suggestions")

        # Анализ поиска
        algolia_result = None
        if algolia_csv or algolia_json:
            print("Analyzing search data...")
            algolia_analyzer = AlgoliaAnalytics()
            if algolia_csv:
                algolia_result = algolia_analyzer.analyze_from_csv(algolia_csv)
            elif algolia_json:
                algolia_result = algolia_analyzer.analyze_from_json(algolia_json)
            print(f"  Found {len(algolia_result.no_results_queries)} no-result queries")

        # Агрегируем
        print("Aggregating results...")
        report = self.aggregate_all(code_result, community_result, algolia_result)
        print(f"  Total gaps: {len(report.gaps)}")

        return report

    def save_to_excel(self, report: AggregatedReport, filename: str = 'doc_gaps_report.xlsx') -> str:
        """
        Сохраняет отчёт в Excel файл.

        Args:
            report: Агрегированный отчёт
            filename: Имя файла

        Returns:
            Путь к созданному файлу
        """
        if not HAS_OPENPYXL:
            print("Warning: openpyxl not installed. Saving to CSV instead.")
            return self.save_to_csv(report, filename.replace('.xlsx', '.csv'))

        filepath = self.output_dir / filename
        wb = Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary, report)

        # Sheet 2: All Gaps
        ws_gaps = wb.create_sheet("Documentation Gaps")
        self._write_gaps_sheet(ws_gaps, report.gaps)

        # Sheet 3: High Priority
        ws_high = wb.create_sheet("High Priority")
        high_priority = [g for g in report.gaps if g.priority == 'high']
        self._write_gaps_sheet(ws_high, high_priority)

        # Sheet 4: By Source
        ws_by_source = wb.create_sheet("By Source")
        self._write_by_source_sheet(ws_by_source, report.gaps)

        wb.save(filepath)
        print(f"Excel report saved: {filepath}")
        return str(filepath)

    def save_to_csv(self, report: AggregatedReport, filename: str = 'doc_gaps_report.csv') -> str:
        """
        Сохраняет отчёт в CSV файл.

        Args:
            report: Агрегированный отчёт
            filename: Имя файла

        Returns:
            Путь к созданному файлу
        """
        filepath = self.output_dir / filename

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([
                'ID', 'Title', 'Description', 'Source', 'Category',
                'Suggested Doc Type', 'Priority', 'Frequency',
                'Keywords', 'Action Required', 'Status'
            ])

            # Data
            for gap in report.gaps:
                writer.writerow([
                    gap.id,
                    gap.title,
                    gap.description,
                    gap.source,
                    gap.category,
                    gap.suggested_doc_type,
                    gap.priority,
                    gap.frequency,
                    ', '.join(gap.keywords),
                    gap.action_required,
                    gap.status,
                ])

        print(f"CSV report saved: {filepath}")
        return str(filepath)

    def save_to_json(self, report: AggregatedReport, filename: str = 'doc_gaps_report.json') -> str:
        """
        Сохраняет отчёт в JSON (для Claude Code анализа).

        Args:
            report: Агрегированный отчёт
            filename: Имя файла

        Returns:
            Путь к созданному файлу
        """
        filepath = self.output_dir / filename

        data = {
            'summary': report.summary,
            'sources_analyzed': report.sources_analyzed,
            'generated_at': report.generated_at,
            'gaps': [asdict(gap) for gap in report.gaps],
        }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(f"JSON report saved: {filepath}")
        return str(filepath)

    def _write_summary_sheet(self, ws, report: AggregatedReport):
        """Записывает Summary sheet."""
        # Стили
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=12)

        ws['A1'] = 'Documentation Gap Analysis Report'
        ws['A1'].font = header_font

        ws['A3'] = 'Generated:'
        ws['B3'] = report.generated_at

        ws['A4'] = 'Sources Analyzed:'
        ws['B4'] = ', '.join(report.sources_analyzed)

        ws['A6'] = 'Summary'
        ws['A6'].font = title_font

        row = 7
        for key, value in report.summary.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            row += 1

        # Колонки auto-width
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _write_gaps_sheet(self, ws, gaps: list[DocumentationGap]):
        """Записывает sheet с гэпами."""
        # Header
        headers = ['ID', 'Title', 'Source', 'Category', 'Doc Type', 'Priority', 'Frequency', 'Action']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        priority_colors = {
            'high': PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid'),
            'medium': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
            'low': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
        }

        for row, gap in enumerate(gaps, 2):
            ws.cell(row=row, column=1, value=gap.id)
            ws.cell(row=row, column=2, value=gap.title)
            ws.cell(row=row, column=3, value=gap.source)
            ws.cell(row=row, column=4, value=gap.category)
            ws.cell(row=row, column=5, value=gap.suggested_doc_type)
            ws.cell(row=row, column=6, value=gap.priority)
            ws.cell(row=row, column=7, value=gap.frequency)
            ws.cell(row=row, column=8, value=gap.action_required)

            # Подсветка по приоритету
            if gap.priority in priority_colors:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = priority_colors[gap.priority]

        # Auto-width (approximate)
        widths = [12, 50, 12, 15, 15, 10, 10, 60]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _write_by_source_sheet(self, ws, gaps: list[DocumentationGap]):
        """Записывает breakdown по источникам."""
        sources = {}
        for gap in gaps:
            if gap.source not in sources:
                sources[gap.source] = []
            sources[gap.source].append(gap)

        row = 1
        for source, source_gaps in sources.items():
            ws.cell(row=row, column=1, value=f'Source: {source.upper()}')
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=row, column=2, value=f'({len(source_gaps)} gaps)')
            row += 2

            for gap in source_gaps[:10]:  # Top 10 per source
                ws.cell(row=row, column=1, value=gap.id)
                ws.cell(row=row, column=2, value=gap.title)
                ws.cell(row=row, column=3, value=gap.priority)
                row += 1

            row += 2

    def _deduplicate_gaps(self, gaps: list[DocumentationGap]) -> list[DocumentationGap]:
        """Объединяет похожие гэпы."""
        # Простая дедупликация по категории + ключевым словам
        seen = {}
        unique_gaps = []

        for gap in gaps:
            key = f"{gap.category}:{gap.suggested_doc_type}"

            if key in seen:
                # Увеличиваем frequency существующего
                existing = seen[key]
                existing.frequency += gap.frequency
                existing.sample_queries.extend(gap.sample_queries)
                existing.keywords = list(set(existing.keywords + gap.keywords))
                # Повышаем приоритет если high
                if gap.priority == 'high':
                    existing.priority = 'high'
            else:
                seen[key] = gap
                unique_gaps.append(gap)

        return unique_gaps

    def _generate_summary(self, report: AggregatedReport) -> dict:
        """Генерирует summary статистику."""
        summary = {
            'total_gaps': len(report.gaps),
            'high_priority': sum(1 for g in report.gaps if g.priority == 'high'),
            'medium_priority': sum(1 for g in report.gaps if g.priority == 'medium'),
            'low_priority': sum(1 for g in report.gaps if g.priority == 'low'),
            'by_source': {},
            'by_doc_type': {},
            'by_category': {},
        }

        for gap in report.gaps:
            summary['by_source'][gap.source] = summary['by_source'].get(gap.source, 0) + 1
            summary['by_doc_type'][gap.suggested_doc_type] = summary['by_doc_type'].get(gap.suggested_doc_type, 0) + 1
            summary['by_category'][gap.category] = summary['by_category'].get(gap.category, 0) + 1

        return summary

    def _map_code_to_doc_type(self, category: str) -> str:
        """Маппит категорию кода на тип документации."""
        mapping = {
            'api_endpoint': 'reference',
            'env_var': 'reference',
            'config_option': 'reference',
            'cli_command': 'reference',
            'public_function': 'reference',
            'breaking_change': 'how-to',  # Migration guide
            'webhook': 'how-to',
            'general': 'how-to',
        }
        return mapping.get(category, 'how-to')

    def _truncate(self, text: str, max_len: int) -> str:
        """Обрезает текст до максимальной длины."""
        if len(text) <= max_len:
            return text
        return text[:max_len - 3] + '...'


if __name__ == '__main__':
    # Demo run
    aggregator = GapAggregator('./reports')

    print("Running full gap analysis...")
    report = aggregator.run_full_analysis(
        repo_path='.',
        since_days=7,
    )

    print(f"\n=== Summary ===")
    print(f"Total gaps found: {report.summary['total_gaps']}")
    print(f"High priority: {report.summary['high_priority']}")
    print(f"By source: {report.summary['by_source']}")
    print(f"By doc type: {report.summary['by_doc_type']}")

    # Сохраняем отчёты
    aggregator.save_to_json(report)
    aggregator.save_to_csv(report)

    if HAS_OPENPYXL:
        aggregator.save_to_excel(report)
